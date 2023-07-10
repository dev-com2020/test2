import openpyxl

b = openpyxl.load_workbook('Data.xlsx')
sht = b.active
c1 = sht.cell(row=3, column=3).value = "Selenium"
b.save('Data.xlsx')
print('Wartość z komórki C3', sht.cell(row=3, column=3).value)